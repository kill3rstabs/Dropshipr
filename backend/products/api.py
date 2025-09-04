from ninja import Router, File
from ninja.files import UploadedFile
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import get_object_or_404
import pandas as pd
import os
import uuid
import json
import csv
from datetime import datetime, timedelta
from django.utils import timezone
from .models import Upload, Product, Scrape
from .utils import ingest_upload, ValidationError, ingest_upload_parallel
from .ebayau_rules import eBayAUBusinessRules
from .amazonau_rules import AmazonAUBusinessRules
from .AmazonAUScrapper import AmazonAUScrapper
from .CostcoAUScrapper import CostcoAUScrapper
from marketplace.models import Marketplace, Store
from vendor.models import Vendor, VendorPrice
from django.db import models

# Scraping imports
import asyncio
import aiohttp
import re
import random
import logging
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Optional, Tuple
from urllib.parse import urlparse
from bs4 import BeautifulSoup
from django.db import transaction
from django.db.models import Q
import pytz
from asgiref.sync import sync_to_async
import threading
from openpyxl import load_workbook
from django.db import close_old_connections
import requests
import io
import subprocess
import sys
from collections import defaultdict

router = Router()

# Configure logging
logger = logging.getLogger(__name__)

# Add more detailed logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('scraper_debug.log')
    ]
)

# Create a specific logger for webhook debugging
webhook_logger = logging.getLogger('webhook_debug')
webhook_logger.setLevel(logging.DEBUG)

# Create a specific logger for database operations
db_logger = logging.getLogger('database_debug')
db_logger.setLevel(logging.DEBUG)

# Async database helper functions
@sync_to_async
def get_ebayau_products_count():
    """Get count of eBayAU products asynchronously"""
    return Product.objects.filter(
        vendor__name__in=eBayAUBusinessRules.EBAYAU_VENDOR_VARIATIONS,
        store__is_active=True
    ).count()

@sync_to_async
def get_ebayau_products():
    """Get eBayAU products asynchronously"""
    return list(Product.objects.filter(
        vendor__name__in=eBayAUBusinessRules.EBAYAU_VENDOR_VARIATIONS,
        store__is_active=True
    ))

@sync_to_async
def get_rescrape_products():
    """Get products that need rescraping asynchronously"""
    return list(Product.objects.filter(
        vendor__name__in=eBayAUBusinessRules.EBAYAU_VENDOR_VARIATIONS,
        scrapes__needs_rescrape=True,
        store__is_active=True
    ))

@sync_to_async
def get_products_by_ids(product_ids):
    """Get products by IDs asynchronously"""
    return list(Product.objects.filter(id__in=product_ids))

# Helper functions for eBayAU SKU deduplication

def _normalize_vendor_sku(sku: str) -> str:
    # Align with URL construction: split off decimal part and trim
    return str(sku).split('.')[0].strip()


def build_vendor_sku_groups(products: List[Product]) -> Tuple[List[Product], Dict[int, List[int]]]:
    """
    Build groups of products sharing the same (vendor_id, normalized vendor_sku).

    Returns:
      - rep_products: list of representative Product objects (one per unique group)
      - rep_to_ids: map of representative product.id -> list of all product IDs in that group
    """
    key_to_ids: Dict[Tuple[int, str], List[int]] = defaultdict(list)
    key_to_rep: Dict[Tuple[int, str], Product] = {}

    for p in products:
        key = (p.vendor_id, _normalize_vendor_sku(p.vendor_sku))
        key_to_ids[key].append(p.id)
        if key not in key_to_rep:
            key_to_rep[key] = p

    rep_products = list(key_to_rep.values())
    rep_to_ids: Dict[int, List[int]] = {rep.id: key_to_ids[key] for key, rep in key_to_rep.items()}
    return rep_products, rep_to_ids

# n8n webhook configuration
N8N_WEBHOOK_URL = os.getenv('N8N_WEBHOOK_URL', 'https://autoecom.wesolucions.com/webhook/ebayau-rescrape')
N8N_WEBHOOK_TIMEOUT = 30  # seconds

async def trigger_n8n_rescrape_webhook(product_ids: List[int], session_id: str) -> bool:
    """
    Trigger n8n webhook for rescraping products
    
    Args:
        product_ids: List of product IDs that need rescraping
        session_id: Current scraping session ID
        
    Returns:
        bool: True if webhook was triggered successfully, False otherwise
    """
    webhook_logger.info(f"=== WEBHOOK TRIGGER START ===")
    webhook_logger.info(f"Session ID: {session_id}")
    webhook_logger.info(f"Product IDs count: {len(product_ids)}")
    webhook_logger.info(f"Product IDs: {product_ids[:10]}...")  # Log first 10 IDs
    webhook_logger.info(f"Webhook URL: {N8N_WEBHOOK_URL}")
    webhook_logger.info(f"Webhook timeout: {N8N_WEBHOOK_TIMEOUT}")
    
    if not product_ids:
        webhook_logger.info("No products need rescraping, skipping n8n webhook")
        return True
    
    webhook_data = {
        "session_id": session_id,
        "product_ids": product_ids,
        "total_products": len(product_ids),
        "triggered_at": timezone.now().isoformat(),
        "source": "ebayau_scraper"
    }
    
    webhook_logger.info(f"Webhook payload: {webhook_data}")
    
    try:
        webhook_logger.info("Creating aiohttp session for webhook call")
        async with aiohttp.ClientSession() as session:
            webhook_logger.info(f"Making POST request to {N8N_WEBHOOK_URL}")
            webhook_logger.info(f"Request headers: {{'Content-Type': 'application/json', 'User-Agent': 'eBayAU-Scraper/1.0'}}")
            
            async with session.post(
                N8N_WEBHOOK_URL,
                json=webhook_data,
                timeout=aiohttp.ClientTimeout(total=N8N_WEBHOOK_TIMEOUT),
                headers={
                    'Content-Type': 'application/json',
                    'User-Agent': 'eBayAU-Scraper/1.0'
                }
            ) as response:
                webhook_logger.info(f"Webhook response status: {response.status}")
                webhook_logger.info(f"Webhook response headers: {dict(response.headers)}")
                
                response_text = await response.text()
                webhook_logger.info(f"Webhook response body: {response_text}")
                
                if response.status == 200:
                    try:
                        response_data = await response.json()
                        webhook_logger.info(f"Webhook JSON response: {response_data}")
                    except:
                        webhook_logger.info("Webhook response is not JSON")
                    
                    webhook_logger.info(f"n8n webhook triggered successfully for {len(product_ids)} products")
                    webhook_logger.info("=== WEBHOOK TRIGGER SUCCESS ===")
                    return True
                else:
                    webhook_logger.error(f"n8n webhook failed with status {response.status}")
                    webhook_logger.error(f"Response: {response_text}")
                    webhook_logger.info("=== WEBHOOK TRIGGER FAILED ===")
                    return False
                    
    except aiohttp.ClientError as e:
        webhook_logger.error(f"aiohttp ClientError calling n8n webhook: {e}")
        webhook_logger.error(f"Error type: {type(e)}")
        webhook_logger.info("=== WEBHOOK TRIGGER FAILED (CLIENT ERROR) ===")
        return False
    except Exception as e:
        webhook_logger.error(f"Unexpected error calling n8n webhook: {e}")
        webhook_logger.error(f"Error type: {type(e)}")
        webhook_logger.error(f"Error traceback: ", exc_info=True)
        webhook_logger.info("=== WEBHOOK TRIGGER FAILED (UNEXPECTED ERROR) ===")
        return False

# Scraping configuration constants
MAX_CONCURRENT_REQUESTS = 5
BATCH_SIZE = 5
TIMEOUT = aiohttp.ClientTimeout(total=45)
RETRY_LIMIT = 1

# Pre-compiled regex patterns for performance
QUANTITY_PATTERN = re.compile(
    r'"NumberValidation","minValue":"(\d+)","maxValue":"(\d+)"'
)
HANDLING_PATTERN = re.compile(
    r'(?<="textSpans":\[\{"_type":"TextSpan","text":"Will usually ship within )[^"]*(?=")'
)
PRICE_CLEAN_PATTERN = re.compile(r'[^\d.]')
STOCK_EXTRACT_PATTERN = re.compile(r'\d+')

# CSS selectors for eBay page elements
SELECTORS = {
    'title': '.x-item-title__mainTitle span',
    'status_message': '.ux-layout-section__textual-display--statusMessage span',
    'price': '.x-price-primary span',
    'seller_away': '.x-alert--ALERT_SA div.ux-message',
    'shipping': '.ux-labels-values--shipping .ux-labels-values__values-content div:nth-of-type(1)',
    'stock': 'div.x-quantity__availability',
    'message': 'div.ux-message__content',
    'select_boxes': 'button.btn--truncated',
    'breadcrumb': '.breadcrumbs li'
}

# Realistic User-Agent strings for rotation
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:128.0) Gecko/20100101 Firefox/128.0',
]

# eBayAU specific constants
EBAYAU_MAX_CONCURRENT_REQUESTS = 50
EBAYAU_BATCH_SIZE = 50
EBAYAU_TIMEOUT = aiohttp.ClientTimeout(total=30)
EBAYAU_RETRY_LIMIT = 3  # Retry 3 times

# eBayAU specific selectors
EBAYAU_SELECTORS = {
    'status_message': '.ux-layout-section__textual-display--statusMessage span',
    'price': '.x-price-primary span',
    'seller_away': '.x-alert--ALERT_SA div.ux-message',
    'shipping': '.ux-labels-values--shipping .ux-labels-values__values-content div:nth-of-type(1)',
    'stock': 'div.x-quantity__availability',
    'stock_fallback': 'div.ux-message',
    'message': 'div.ux-message__content',
    'select_boxes': 'button.btn--truncated',
    'specific_error_header': 'p.error-header-v2__title'
}

# eBayAU regex patterns
EBAYAU_QUANTITY_PATTERN = re.compile(
    r'"NumberValidation","minValue":"(\d+)","maxValue":"(\d+)"'
)
EBAYAU_HANDLING_PATTERN = re.compile(
    r'(?<="textSpans":\[\{"_type":"TextSpan","text":"Will usually (?:post|ship) within )[^"]*(?=")'
)

# eBayAU cookies (exact same as your script)
EBAYAU_COOKIES = {
    's': 'CgAD4ACBoZ66CY2ZkYjUzMTAxOTcwYTQ3ZTg3NzFkNTU2ZmZmNmUyNGIyMJaY',
    '__uzma': 'fec443b7-de63-449d-a449-704926aaeb69',
    '__uzmb': '1751538946',
    '__uzme': '9985',
    'ak_bmsc': '1ACBBD284EE8CF0470D052F667891336~000000000000000000000000000000~YAAQ5CE1F4nQbMaXAQAAqFXbzxyyRQ21qGWJKt0BpuNhDJ/W/pJ6pzaF0p3d+uIGjR2wRRaW9u8IGdyZNxIsVCkE+z7e+umCf2i0xH9TAKqMvOFWQ9fxEV7QNSTSYI0hRhHa2pxsPuItThD83kAOKaVftxzeIq8epwGjAQgyCS2PMqKTdid0FXlsC7WXMbaB0xRjNCb0Ar0vLwD+2IDH85F8GLl6C1Tyg9HzdaieMnJJxY4HXhWEOQyW/Wjdmaz/tAPXuGceCF8l8LoYtbdZnJaizYBtT0iFLh8DX8cQNax8EIjs3/5+aJbHu2boJsqZ9A/sJ4QgnHSgQPUDv163J7IELq2w97cp7hzZbC8rCcQk8zitfjh0jt/SAo1YdnnQEgFK/X0tEQ==',
    '__ssds': '2',
    '__uzjsr2': 'a9be0cd8e',
    'ebay': '%5Ejs%3D1%5Esbf%3D%23000000%5E',
    '__uzmc': '916572543451',
    '__uzmd': '1751538964',
    '__uzmf': '7f600098f91af7-6075-4051-b45b-0cb4eccb3a0e175153894674818229-517ed63e82430b3025',
    'bm_sv': '882C7527777792580D28A66D006AEA7A~YAAQHig0F2IsdcaXAQAAyprbzxzRDvs0Unf96hCW8oNBB9ngSKAcXb/RdVdB2I+Lf0S8p1moZNJQqANEiryeeIF/QtnXDJ5IPxQbC/5g2bKqxmuIgBOVo/N/qrsfb/3oj7sODDfn/pAQ5BiqdT5r2S5Yi2r/afYxxZK4cRLc03BLOHTIZ2i0JPatqnUlqkSnCs6P9BMzgYCKcRlcoTPdGooKCs0u2AXIcFIrYA14rz+qN6cp+Vb6aEqC+4mr~1',
    'dp1': 'bpbf/%234000000000000000000006a479096^bl/PK6c28c416^',
    'nonsession': 'BAQAAAZZay1gcAAaAADMACGpHkJYyNjAxLEFVUwDKACBsKMQWY2ZkYjUzMTAxOTcwYTQ3ZTg3NzFkNTU2ZmZmNmUyNGIAywABaGZkHjjqkHe4DUZuXtDGYqfTKa7/GdfXiQ**',
    '__deba': '_DrPrcvjLRsDEPlDcpZmdKWQXXLk9lguZC_6F7S7H8K4eofYDZrnCYcK7BuGjkdhXR8UW1J2_XMwSXzXGqPFGDx58wBA8Pdbv7F_V-WNcj7FdXNkj_vv2YWRlCdh5jdow2a1n18xz_QaS8tcNp9ysw=='
}

# Scraping Helper Functions
def validate_ebay_item_number(vendor_sku: str) -> tuple[bool, str]:
    """Validate eBay item number format."""
    try:
        item_num = str(vendor_sku).split('.')[0]
        
        if not item_num.isdigit():
            return False, "Not numeric"
        
        if len(item_num) < 10 or len(item_num) > 12:
            return False, f"Invalid length: {len(item_num)}"
            
        return True, "Valid"
    except Exception as e:
        logger.warning(f"Error validating eBay item number {vendor_sku}: {e}")
        return False, "Invalid format"

def parse_price_to_decimal(price_text: str) -> Optional[Decimal]:
    """Parse price text to decimal with 2 decimal places."""
    if not price_text:
        return None
    
    try:
        clean_price = PRICE_CLEAN_PATTERN.sub('', str(price_text))
        if clean_price:
            return Decimal(clean_price).quantize(Decimal('0.01'))
        return None
    except (InvalidOperation, ValueError) as e:
        logger.warning(f"Error parsing price '{price_text}': {e}")
        return None

# eBayAU Helper Functions
def get_ebayau_product_quantity(soup):
    """Extract quantity information from eBayAU page source."""
    view_page_source = str(soup)
    match = EBAYAU_QUANTITY_PATTERN.search(view_page_source)
    if match:
        min_value = match.group(1)
        max_value = match.group(2)
        return f"Min: {min_value}, Max: {max_value}"
    else:
        return "Quantity info not found"

def get_ebayau_ended_listings(soup):
    """Extract ended listings status from eBayAU page."""
    status_message_element = soup.select_one(EBAYAU_SELECTORS['status_message'])
    return status_message_element.get_text(strip=True) if status_message_element else ""

def get_ebayau_product_price(soup):
    """Extract product price from eBayAU page."""
    price_element = soup.select_one(EBAYAU_SELECTORS['price'])
    if price_element:
        return price_element.get_text(strip=True)
    else:
        return None

def get_ebayau_seller_away(soup):
    """Extract seller away message from eBayAU page."""
    seller_away_element = soup.select_one(EBAYAU_SELECTORS['seller_away'])
    return seller_away_element.get_text(strip=True) if seller_away_element else ""

def get_ebayau_shipping_info(soup):
    """Extract shipping information from eBayAU page."""
    shipping_element = soup.select_one(EBAYAU_SELECTORS['shipping'])
    return shipping_element.get_text(strip=True) if shipping_element else "No shipping info"

def get_ebayau_handling_time(soup):
    """Extract handling time from eBayAU page source."""
    view_page_source = str(soup)
    match = EBAYAU_HANDLING_PATTERN.search(view_page_source)
    if match:
        full_message = f"Will usually post/ship within {match.group()}"
        return full_message
    else:
        return "Handling time info not found"

def parse_ebayau_product_details_from_soup(soup, url):
    """Parse all product details from eBayAU page soup."""
    stock_element = soup.select_one(EBAYAU_SELECTORS['stock'])
    if not stock_element:
        stock_element = soup.select_one(EBAYAU_SELECTORS['stock_fallback'])
    
    quantity = get_ebayau_product_quantity(soup)
    price = get_ebayau_product_price(soup)
    ended_listings = get_ebayau_ended_listings(soup)
    seller_away = get_ebayau_seller_away(soup)
    shipping_info = get_ebayau_shipping_info(soup)
    handling_time = get_ebayau_handling_time(soup)
    
    stock = None
    if stock_element:
        stock = stock_element.get_text(strip=True)
    
    parsed_url = urlparse(url)
    ebay_item_number = parsed_url.path.split('/')[-1].split('?')[0]
    
    return {
        'ebay_item_number': ebay_item_number,
        'quantity': quantity,
        'stock': stock,
        'price': price,
        'ended_listings': ended_listings,
        'seller_away': seller_away,
        'shipping_info': shipping_info,
        'handling_time': handling_time
    }

def parse_stock_to_int(stock_text: str) -> Optional[int]:
    """Parse stock text to integer."""
    if not stock_text:
        return None
    
    try:
        numbers = STOCK_EXTRACT_PATTERN.findall(str(stock_text))
        if numbers:
            return int(numbers[0])
        return None
    except (ValueError, IndexError) as e:
        logger.warning(f"Error parsing stock '{stock_text}': {e}")
        return None

def generate_error_log_filename(session_id: str) -> str:
    """Generate filename for error logs."""
    return f"scrapping_logs_{session_id}.xlsx"

def create_error_log_excel(results: List[Dict[str, Any]], session_id: str) -> str:
    """Create Excel error log file."""
    try:
        filename = generate_error_log_filename(session_id)
        filepath = os.path.join("uploads", filename)
        
        # Prepare data for Excel
        excel_data = []
        for result in results:
            try:
                product_id = result.get('product_id')
                if product_id:
                    product = Product.objects.select_related('vendor', 'marketplace', 'store', 'upload').get(id=product_id)
                    
                    # Determine status
                    status = "SUCCESS" if result.get('success') else "FAILED"
                    response_text = str(result) if result.get('success') else result.get('error_status', 'Unknown error')
                    
                    excel_data.append({
                        'Vendor Name': product.vendor.name,
                        'Vendor ID': product.vendor_sku,
                        'Is Variation': 'Yes' if product.variation_id else 'No',
                        'Variation ID': product.variation_id,
                        'Marketplace Name': product.marketplace.name,
                        'Store Name': product.store.name,
                        'Marketplace Parent SKU': product.marketplace_parent_sku,
                        'Marketplace Child SKU': product.marketplace_child_sku,
                        'Marketplace ID': '',  # Not stored in our model
                        'Product ID': product.id,
                        'STATUS': status,
                        'Response from scrapper': response_text
                    })
            except Product.DoesNotExist:
                logger.error(f"Product {product_id} not found for error log")
            except Exception as e:
                logger.error(f"Error processing product {product_id} for error log: {e}")
        
        # Create DataFrame and save to Excel
        df = pd.DataFrame(excel_data)
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        logger.info(f"Error log created: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"Error creating error log: {e}")
        return ""

def _quick_file_info(file_path: str, file_extension: str):
    """Fast row count and first-row metadata without loading entire dataset"""
    vendor_name = "Unknown"
    marketplace_name = "Unknown"
    if file_extension == '.csv':
        try:
            with open(file_path, 'rb') as f:
                total_lines = sum(1 for _ in f)
            items_uploaded = max(total_lines - 1, 0)
        except Exception:
            items_uploaded = 0
        try:
            import csv as _csv
            with open(file_path, 'r', encoding='utf-8', newline='') as f:
                reader = _csv.DictReader(f)
                first = next(reader, None)
                if first:
                    vendor_name = first.get('Vendor Name') or "Unknown"
                    marketplace_name = first.get('Marketplace Name') or "Unknown"
        except Exception:
            pass
        return items_uploaded, vendor_name, marketplace_name
    else:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            max_row = ws.max_row or 1
            items_uploaded = max(max_row - 1, 0)
            headers = [cell.value for cell in ws[1]]
            second = [cell.value for cell in ws[2]] if max_row >= 2 else []
            header_index = {h: i for i, h in enumerate(headers) if h}
            if second:
                if 'Vendor Name' in header_index:
                    v = second[header_index['Vendor Name']]
                    vendor_name = str(v) if v is not None else "Unknown"
                if 'Marketplace Name' in header_index:
                    m = second[header_index['Marketplace Name']]
                    marketplace_name = str(m) if m is not None else "Unknown"
        except Exception:
            items_uploaded = 0
        return items_uploaded, vendor_name, marketplace_name


def _progress_file_path(upload_id: int) -> str:
    uploads_dir = os.path.join("uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    return os.path.join(uploads_dir, f"progress_{upload_id}.json")


def _read_progress(upload_id: int):
    path = _progress_file_path(upload_id)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None

# Email webhook configuration
EMAIL_WEBHOOK_URL = os.getenv('EMAIL_WEBHOOK_URL', 'https://autoecom.wesolucions.com/webhook/send-email')
EMAIL_WEBHOOK_TIMEOUT = 30  # seconds

def generate_system_products_csv() -> str:
    """
    Generate system_products.csv file and return the file path
    """
    try:
        # Get all products with related data including vendor prices
        products = Product.objects.select_related('vendor', 'marketplace', 'store').prefetch_related('latest_price').all()
        
        # Create uploads directory if it doesn't exist
        uploads_dir = os.path.join("uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Generate unique filename with timestamp
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"system_products_{timestamp}.csv"
        csv_path = os.path.join(uploads_dir, csv_filename)
        
        # Write CSV file
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write headers
            headers = [
                'Vendor Name',
                'Vendor ID',
                'Is Variation',
                'Variation ID', 
                'Marketplace Name',
                'Store Name',
                'Marketplace Parent SKU',
                'Marketplace Child SKU',
                'Marketplace ID',
                'Vendor Price',
                'Vendor Inventory'
            ]
            writer.writerow(headers)
            
            # Write product data
            for product in products:
                vendor_price = product.latest_price
                
                row = [
                    product.vendor.name if product.vendor else '',
                    product.vendor_sku or '',
                    'Yes' if product.variation_id else 'No',
                    product.variation_id or '',
                    product.marketplace.name if product.marketplace else '',
                    product.store.name if product.store else '',
                    product.marketplace_parent_sku or '',
                    product.marketplace_child_sku or '',
                    product.marketplace_external_id or '',
                    vendor_price.price if vendor_price and vendor_price.price else '',
                    vendor_price.stock if vendor_price and vendor_price.stock else '0'
                ]
                writer.writerow(row)
        
        logger.info(f"Generated system products CSV: {csv_path}")
        return csv_path
        
    except Exception as e:
        logger.error(f"Error generating system products CSV: {e}")
        return ""

def build_system_products_csv_bytes() -> bytes:
    """Generate the system products CSV in-memory (bytes), matching the /export endpoint."""
    products = Product.objects.select_related('vendor', 'marketplace', 'store').prefetch_related('latest_price').all()
    import io as _io
    import csv as _csv
    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    headers = [
        'Vendor Name', 'Vendor ID', 'Is Variation', 'Variation ID',
        'Marketplace Name', 'Store Name', 'Marketplace Parent SKU',
        'Marketplace Child SKU', 'Marketplace ID', 'Vendor Price', 'Vendor Inventory'
    ]
    writer.writerow(headers)
    for product in products:
        vendor_price = product.latest_price
        row = [
            product.vendor.name if product.vendor else '',
            product.vendor_sku or '',
            'Yes' if product.variation_id else 'No',
            product.variation_id or '',
            product.marketplace.name if product.marketplace else '',
            product.store.name if product.store else '',
            product.marketplace_parent_sku or '',
            product.marketplace_child_sku or '',
            product.marketplace_external_id or '',
            vendor_price.price if vendor_price and vendor_price.price else '',
            vendor_price.stock if vendor_price and vendor_price.stock else '0',
        ]
        writer.writerow(row)
    return buffer.getvalue().encode('utf-8')

def send_scraping_complete_email(session_id: str, scraping_stats: dict, csv_file_path: str = None, recipient_email: str = None):
    """
    Send scraping completion notification email with CSV attachment.
    Always send multipart/form-data with a 'data' file so n8n Gmail node receives a binary.
    If no CSV exists on disk, attach a freshly generated in-memory CSV.
    """
    try:
        if not recipient_email:
            recipient_email = os.getenv('DEFAULT_NOTIFICATION_EMAIL', 'afraaz.prettyandpractical@gmail.com')
        
        # HTML email template for scraping completion
        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Scraping Complete - AutoEcomm</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f4f4f4; }
        .container { background: white; padding: 30px; border-radius: 10px; box-shadow: 0 0 20px rgba(0,0,0,0.1); }
        .header { text-align: center; border-bottom: 3px solid #28a745; padding-bottom: 20px; margin-bottom: 30px; }
        .logo { font-size: 28px; font-weight: bold; color: #28a745; margin-bottom: 10px; }
        .status-success { background: #d4edda; color: #155724; padding: 15px; border-radius: 5px; border-left: 4px solid #28a745; margin: 20px 0; }
        .details { background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 20px 0; }
        .details h3 { margin-top: 0; color: #495057; }
        .detail-row { display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #dee2e6; }
        .detail-row:last-child { border-bottom: none; }
        .detail-label { font-weight: 600; color: #6c757d; }
        .detail-value { color: #495057; }
        .footer { text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #dee2e6; color: #6c757d; font-size: 14px; }
        .attachment-info { background: #e7f3ff; border: 1px solid #b3d9ff; padding: 15px; border-radius: 5px; margin: 15px 0; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="logo">AutoEcomm</div>
            <h2>Scraping Job Complete</h2>
        </div>

        <div class="status-success">
            <strong>Success!</strong> Your scraping job has been completed successfully.
        </div>

        <div class="details">
            <h3>Scraping Statistics</h3>
            <div class="detail-row"><span class="detail-label">Session ID:</span><span class="detail-value">{SESSION_ID}</span></div>
            <div class="detail-row"><span class="detail-label">Total Products:</span><span class="detail-value">{TOTAL_PRODUCTS}</span></div>
            <div class="detail-row"><span class="detail-label">Successful Scrapes:</span><span class="detail-value">{SUCCESSFUL_SCRAPES}</span></div>
            <div class="detail-row"><span class="detail-label">Failed Scrapes:</span><span class="detail-value">{FAILED_SCRAPES}</span></div>
            <div class="detail-row"><span class="detail-label">Success Rate:</span><span class="detail-value">{SUCCESS_RATE}%</span></div>
            <div class="detail-row"><span class="detail-label">Duration:</span><span class="detail-value">{DURATION}</span></div>
        </div>

        {ATTACHMENT_SECTION}

        <div class="footer"><p>This is an automated notification from AutoEcomm.</p><p>If you have any questions, please contact support.</p></div>
    </div>
</body>
</html>
        """
        
        has_real_csv = bool(csv_file_path and os.path.isfile(csv_file_path))
        
        # Determine attachment section
        if has_real_csv:
            attachment_section = '''
        <div class="attachment-info">
            <strong>Attachment:</strong> System Products CSV is attached to this email.
        </div>'''
        else:
            attachment_section = ''
        
        # Replace placeholders
        html_message = html_template.replace('{SESSION_ID}', session_id)
        html_message = html_message.replace('{TOTAL_PRODUCTS}', str(scraping_stats.get('total_products', 0)))
        html_message = html_message.replace('{SUCCESSFUL_SCRAPES}', str(scraping_stats.get('successful_scrapes', 0)))
        html_message = html_message.replace('{FAILED_SCRAPES}', str(scraping_stats.get('failed_scrapes', 0)))
        html_message = html_message.replace('{SUCCESS_RATE}', f"{scraping_stats.get('success_rate', 0):.1f}")
        html_message = html_message.replace('{DURATION}', str(scraping_stats.get('duration', 'Unknown')))
        html_message = html_message.replace('{ATTACHMENT_SECTION}', attachment_section)
        
        # Form fields
        email_payload = {
            "to": recipient_email,
            "subject": f"Scraping Complete - Session {session_id}",
            "message": html_message,
        }
        
        # Always send a 'data' file for n8n Gmail node
        try:
            if has_real_csv:
                filename = os.path.basename(csv_file_path)
                with open(csv_file_path, 'rb') as f:
                    files = {"data": (filename, f, "text/csv")}
                    response = requests.post(
                        EMAIL_WEBHOOK_URL,
                        data=email_payload,
                        files=files,
                        timeout=EMAIL_WEBHOOK_TIMEOUT,
                    )
            else:
                # Build the same system products CSV in-memory
                csv_bytes = build_system_products_csv_bytes()
                filename = f"system_products_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
                mem = io.BytesIO(csv_bytes if csv_bytes else b"note,No rows\n")
                files = {"data": (filename, mem, "text/csv")}
                response = requests.post(
                    EMAIL_WEBHOOK_URL,
                    data=email_payload,
                    files=files,
                    timeout=EMAIL_WEBHOOK_TIMEOUT,
                )
            
            response.raise_for_status()
            logger.info(f"Scraping completion email sent successfully to {recipient_email}")
            return True
        except Exception as e:
            logger.error(f"Failed to send scraping email: {e}")
            return False
    
    except Exception as e:
        logger.exception(f"Error sending scraping completion email: {e}")
        return False

def send_upload_notification_email(upload_data: dict, recipient_email: str = None):
    """
    Send upload completion notification email using webhook API
    
    Args:
        upload_data: Dictionary containing upload details
        recipient_email: Email address to send to (optional)
    """
    try:
        if not recipient_email:
            # Default email - you can get this from user profile or settings
            recipient_email = os.getenv('DEFAULT_NOTIFICATION_EMAIL', 'afraaz.prettyandpractical@gmail.com')
        
        # Read HTML template from file
        template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'email_upload_complete.html')
        with open(template_path, 'r', encoding='utf-8') as f:
            html_template = f.read()
        
        # Determine status section
        status = upload_data.get('status', 'unknown').lower()
        if status == 'completed':
            status_section = '''
        <div class="status-success">
            <strong>✅ Success!</strong> Your file has been processed successfully.
        </div>'''
        else:
            status_section = '''
        <div class="status-failed">
            <strong>❌ Failed!</strong> There was an issue processing your file.
        </div>'''
        
        # Determine error section
        error_logs = upload_data.get('error_logs', '')
        if error_logs and error_logs != 'No errors':
            error_section = f'''
        <div class="error-logs">
            <strong>Error Details:</strong><br>
            {error_logs}
        </div>'''
        else:
            error_section = ''
        
        # Replace placeholders
        html_message = html_template.replace('{STATUS_SECTION}', status_section)
        html_message = html_message.replace('{ERROR_SECTION}', error_section)
        html_message = html_message.replace('{FILE_NAME}', upload_data.get('file_name', 'Unknown'))
        html_message = html_message.replace('{UPLOAD_DATE}', upload_data.get('upload_date', 'Unknown'))
        html_message = html_message.replace('{VENDOR_NAME}', upload_data.get('vendor_name', 'Unknown'))
        html_message = html_message.replace('{MARKETPLACE_NAME}', upload_data.get('marketplace_name', 'Unknown'))
        html_message = html_message.replace('{ITEMS_UPLOADED}', str(upload_data.get('items_uploaded', 0)))
        html_message = html_message.replace('{ITEMS_ADDED}', str(upload_data.get('items_added', 0)))
        html_message = html_message.replace('{STATUS}', status.title())
        
        # Prepare email payload
        email_payload = {
            "to": recipient_email,
            "subject": f"Upload {upload_data.get('status', 'Complete').title()} - {upload_data.get('file_name', 'File')}",
            "message": html_message
        }
        
        # Send email via webhook
        response = requests.post(
            EMAIL_WEBHOOK_URL,
            json=email_payload,
            timeout=EMAIL_WEBHOOK_TIMEOUT,
            headers={'Content-Type': 'application/json'}
        )
        
        if response.status_code == 200:
            logger.info(f"Upload notification email sent successfully to {recipient_email}")
            return True
        else:
            logger.error(f"Failed to send email. Status: {response.status_code}, Response: {response.text}")
            return False
            
    except Exception as e:
        logger.exception(f"Error sending upload notification email: {e}")
        return False

def _process_upload_in_background(upload_id: int):
    close_old_connections()
    email_data = {}
    
    try:
        logger.info(f"Starting background ingest for upload_id={upload_id}")
        
        # Get upload info for email
        upload = Upload.objects.get(id=upload_id)
        try:
            info = json.loads(upload.note) if upload.note else {}
        except Exception:
            info = {}
        
        email_data = {
            'file_name': upload.original_name,
            'upload_date': upload.expires_at.strftime("%Y-%m-%d %H:%M"),
            'vendor_name': info.get('vendorName', 'Unknown'),
            'marketplace_name': info.get('marketplace', 'Unknown'),
            'items_uploaded': info.get('itemsUploaded', 0),
        }
        
        # Use parallel ingestion for performance on large files
        processed_count = ingest_upload_parallel(upload_id, workers=4, batch_size=500)

        info.update({
            'status': 'completed',
            'itemsAdded': processed_count,
            'errorLogs': 'No errors',
            'itemsProcessed': processed_count
        })
        upload.note = json.dumps(info)
        upload.save(update_fields=['note'])
        
        # Update email data for success notification
        email_data.update({
            'items_added': processed_count,
            'status': 'completed',
            'error_logs': 'No errors'
        })
        
        logger.info(f"Completed background ingest for upload_id={upload_id} itemsAdded={processed_count}")
        
    except ValidationError as e:
        upload = Upload.objects.filter(id=upload_id).first()
        if upload:
            try:
                info = json.loads(upload.note) if upload.note else {}
            except Exception:
                info = {}
            info.update({
                'status': 'failed',
                'itemsAdded': 0,
                'errorLogs': str(e),
                'errorType': e.error_type
            })
            upload.note = json.dumps(info)
            upload.save(update_fields=['note'])
            
            # Update email data for failure notification
            email_data.update({
                'items_added': 0,
                'status': 'failed',
                'error_logs': str(e)
            })
            
        logger.exception(f"Validation error during background ingest upload_id={upload_id}: {e}")
        
    except Exception as e:
        upload = Upload.objects.filter(id=upload_id).first()
        if upload:
            try:
                info = json.loads(upload.note) if upload.note else {}
            except Exception:
                info = {}
            info.update({
                'status': 'failed',
                'itemsAdded': 0,
                'errorLogs': str(e)
            })
            upload.note = json.dumps(info)
            upload.save(update_fields=['note'])
            
            # Update email data for failure notification
            email_data.update({
                'items_added': 0,
                'status': 'failed',
                'error_logs': str(e)
            })
            
        logger.exception(f"Unexpected error during background ingest upload_id={upload_id}: {e}")
        
    finally:
        # Send notification email
        if email_data:
            try:
                send_upload_notification_email(email_data)
            except Exception as email_error:
                logger.error(f"Failed to send notification email for upload_id={upload_id}: {email_error}")
        
        close_old_connections()

@router.post("/upload/")
def upload_file(request, file: UploadedFile = File(...)):
    """
    Handle file upload for product mapping with comprehensive validation
    """
    try:
        # Validate file type
        allowed_extensions = ['.csv', '.xlsx', '.xls']
        file_extension = os.path.splitext(file.name)[1].lower()
        
        if file_extension not in allowed_extensions:
            return {
                "success": False, 
                "error": "Invalid file type. Please upload CSV or Excel files only.",
                "errorType": "INVALID_FILE_TYPE"
            }
        
        # Generate unique filename
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        uploads_dir = os.path.join("uploads")
        file_path = os.path.join(uploads_dir, unique_filename)
        
        # Ensure uploads directory exists
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Save file
        with open(file_path, "wb") as f:
            for chunk in file.chunks():
                f.write(chunk)
        
        # Create Upload record
        upload = Upload.objects.create(
            original_name=file.name,
            stored_key=file_path,
            expires_at=timezone.now() + timedelta(days=30)  # Keep for 30 days
        )
        
        # Get basic file info for response (fast path)
        try:
            items_uploaded, vendor_name, marketplace_name = _quick_file_info(file_path, file_extension)
        except Exception as e:
            # Clean up uploaded file if we can't parse it
            if os.path.exists(file_path):
                os.remove(file_path)
            upload.delete()
            return {
                "success": False, 
                "error": f"File parsing error: {str(e)}",
                "errorType": "FILE_PARSING_ERROR"
            }
        
        # Initialize status and persist for polling
        status_info = {
            'status': 'processing',
            'vendorName': vendor_name,
            'marketplace': marketplace_name,
            'itemsUploaded': items_uploaded,
            'itemsAdded': 0,
            'errorLogs': 'Processing',
            'itemsProcessed': 0,
            'totalItems': items_uploaded
        }
        upload.note = json.dumps(status_info)
        upload.save(update_fields=['note'])

        # Fire-and-forget background processing
        threading.Thread(
            target=_process_upload_in_background,
            args=(upload.id,),
            daemon=True,
        ).start()
        logger.info(f"Spawned background ingest thread for upload_id={upload.id}")

        return {
            "success": True,
            "upload_id": upload.id,
            "date": timezone.now().strftime("%Y-%m-%d"),
            "userName": request.user.username if hasattr(request, 'user') and request.user.is_authenticated else "System",
            "vendorName": vendor_name,
            "marketplace": marketplace_name,
            "itemsUploaded": items_uploaded,
            "itemsAdded": 0,
            "status": "processing",
            "errorLogs": "Processing"
        }
        
    except Exception as e:
        return {
            "success": False, 
            "error": f"Upload failed: {str(e)}",
            "errorType": "UPLOAD_ERROR"
        }

@router.get("/uploads/")
def get_uploads(request, page: int = 1, page_size: int = 10):
    """
    Get upload records with pagination and preserved status
    """
    try:
        # Validate pagination parameters
        page = max(1, page)
        page_size = min(max(1, page_size), 100)  # Max 100 items per page
        
        # Calculate offset
        offset = (page - 1) * page_size
        
        # Get total count
        total_count = Upload.objects.count()
        
        # Get paginated uploads
        uploads = Upload.objects.all().order_by('-id')[offset:offset + page_size]
        
        results = []
        for upload in uploads:
            try:
                # Check if we already have stored status information
                stored_info = None
                if upload.note:
                    try:
                        stored_info = json.loads(upload.note)
                        # Validate it's our status info format
                        if not isinstance(stored_info, dict) or 'status' not in stored_info:
                            stored_info = None
                    except (json.JSONDecodeError, TypeError):
                        stored_info = None
                
                # If we have stored info, use it (preserve status)
                if stored_info:
                    vendor_name = stored_info.get('vendorName', 'Unknown')
                    marketplace_name = stored_info.get('marketplace', 'Unknown')
                    items_uploaded = stored_info.get('itemsUploaded', 0)
                    items_added = stored_info.get('itemsAdded', 0)
                    status = stored_info.get('status', 'pending')
                    error_logs = stored_info.get('errorLogs', 'No errors')
                else:
                    # First time processing - calculate status and store it
                    # Try to read the file to get details
                    if upload.stored_key.endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(upload.stored_key)
                    else:
                        df = pd.read_csv(upload.stored_key)
                    
                    items_uploaded = len(df)
                    
                    # Get vendor and marketplace names from first row
                    vendor_name = df['Vendor Name'].iloc[0] if len(df) > 0 and 'Vendor Name' in df.columns else "Unknown"
                    marketplace_name = df['Marketplace Name'].iloc[0] if len(df) > 0 and 'Marketplace Name' in df.columns else "Unknown"
                    
                    # Try to determine status by checking if products were created
                    try:
                        vendor = Vendor.objects.filter(code=vendor_name).first()
                        marketplace = Marketplace.objects.filter(code=marketplace_name).first()
                        
                        if vendor and marketplace:
                            # Count products created from this upload (approximate)
                            items_added = Product.objects.filter(vendor=vendor, marketplace=marketplace).count()
                            status = "completed" if items_added > 0 else "pending"
                            error_logs = "No errors"
                        else:
                            items_added = 0
                            status = "failed"
                            error_logs = "Vendor or marketplace not found"
                            
                    except Exception as e:
                        items_added = 0
                        status = "failed"
                        error_logs = f"Status check error: {str(e)}"
                    
                    # Store the calculated information for future requests
                    status_info = {
                        'status': status,
                        'vendorName': vendor_name,
                        'marketplace': marketplace_name,
                        'itemsUploaded': items_uploaded,
                        'itemsAdded': items_added,
                        'errorLogs': error_logs
                    }
                    
                    # Save to upload note field
                    upload.note = json.dumps(status_info)
                    upload.save()
                    
            except Exception as e:
                # Fallback for any errors
                vendor_name = "Unknown"
                marketplace_name = "Unknown"
                items_uploaded = 0
                items_added = 0
                status = "failed"
                error_logs = f"File read error: {str(e)}"
            
            results.append({
                "id": upload.id,
                "date": upload.expires_at.strftime("%Y-%m-%d"),
                "userName": "System",  # You can add user tracking later
                "vendorName": vendor_name,
                "marketplace": marketplace_name,
                "itemsUploaded": items_uploaded,
                "itemsAdded": items_added,
                "status": status,
                "errorLogs": error_logs
            })
        
        # Calculate pagination metadata
        total_pages = (total_count + page_size - 1) // page_size
        has_next = page < total_pages
        has_prev = page > 1
        
        return {
            "success": True, 
            "uploads": results,
            "pagination": {
                "current_page": page,
                "page_size": page_size,
                "total_count": total_count,
                "total_pages": total_pages,
                "has_next": has_next,
                "has_prev": has_prev
            }
        }
        
    except Exception as e:
        return {"success": False, "error": str(e), "uploads": []}

@router.delete("/upload/{upload_id}")
def delete_upload(request, upload_id: int):
    """
    Delete an upload and all products created by it
    """
    try:
        # Get the upload record
        try:
            upload = Upload.objects.get(id=upload_id)
        except Upload.DoesNotExist:
            return {
                "success": False,
                "error": "Upload not found",
                "errorType": "UPLOAD_NOT_FOUND"
            }
        
        deletion_summary = {
            "products_deleted": 0,
            "vendor_prices_deleted": 0,
            "file_deleted": False,
            "upload_deleted": False
        }
        
        errors = []
        
        # Get upload details for product identification
        try:
            # Read the file to get vendor/marketplace info
            if upload.stored_key.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(upload.stored_key)
            else:
                df = pd.read_csv(upload.stored_key)
            
            if len(df) > 0:
                # Find and delete products created by this upload
                from django.db.models import Q
                from marketplace.models import Store
                
                products_to_delete = []
                
                # Get exact product combinations from the uploaded file
                for index, row in df.iterrows():
                    vendor_name = str(row['Vendor Name']).strip()
                    marketplace_name = str(row['Marketplace Name']).strip()
                    store_name = str(row['Store Name']).strip()
                    marketplace_child_sku = str(row['Marketplace Child SKU']).strip()
                    vendor_sku = str(row['Vendor ID']).strip() if 'Vendor ID' in row else ''
                    
                    # Find the exact vendor, marketplace, and store
                    vendor = Vendor.objects.filter(
                        Q(code=vendor_name) | Q(name=vendor_name)
                    ).first()
                    
                    marketplace = Marketplace.objects.filter(
                        Q(code=marketplace_name) | Q(name=marketplace_name)
                    ).first()
                    
                    store = Store.objects.filter(
                        name=store_name,
                        marketplace=marketplace
                    ).first()
                    
                    if vendor and marketplace and store:
                        # Find the exact product that matches this row from the file
                        matching_product = Product.objects.filter(
                            vendor=vendor,
                            marketplace=marketplace,
                            store=store,
                            marketplace_child_sku=marketplace_child_sku,
                            vendor_sku=vendor_sku
                        ).first()
                        
                        if matching_product and matching_product not in products_to_delete:
                            products_to_delete.append(matching_product)
                
                # Delete VendorPrice records first (they reference products)
                from django.db import transaction
                
                with transaction.atomic():
                    vendor_prices_count = 0
                    products_count = 0
                    
                    for product in products_to_delete:
                        # Delete associated VendorPrice records
                        vendor_price_count = VendorPrice.objects.filter(product=product).count()
                        VendorPrice.objects.filter(product=product).delete()
                        vendor_prices_count += vendor_price_count
                    
                    # Delete the products
                    products_count = len(products_to_delete)
                    for product in products_to_delete:
                        product.delete()
                    
                    deletion_summary["products_deleted"] = products_count
                    deletion_summary["vendor_prices_deleted"] = vendor_prices_count
                    
        except Exception as e:
            errors.append(f"Error processing products: {str(e)}")
        
        # Delete the physical file
        try:
            if os.path.exists(upload.stored_key):
                os.remove(upload.stored_key)
                deletion_summary["file_deleted"] = True
            else:
                errors.append("Physical file not found or already deleted")
        except Exception as e:
            errors.append(f"Error deleting file: {str(e)}")
        
        # Delete the upload record
        try:
            upload.delete()
            deletion_summary["upload_deleted"] = True
        except Exception as e:
            errors.append(f"Error deleting upload record: {str(e)}")
        
        # Determine response based on success/failures
        if errors:
            return {
                "success": True,  # Partial success
                "message": "Upload deletion completed with some issues",
                "summary": deletion_summary,
                "warnings": errors,
                "partial": True
            }
        else:
            return {
                "success": True,
                "message": f"Upload deleted successfully. Removed {deletion_summary['products_deleted']} products and {deletion_summary['vendor_prices_deleted']} vendor prices.",
                "summary": deletion_summary,
                "partial": False
            }
            
    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to delete upload: {str(e)}",
            "errorType": "DELETE_ERROR"
        }

@router.post("/bulk-delete/")
def bulk_delete_products(request, file: UploadedFile = File(...)):
    """
    Handle bulk deletion of products based on CSV file with Child SKU and Store Name
    """
    try:
        # Validate file type
        allowed_extensions = ['.csv', '.xlsx', '.xls']
        file_extension = os.path.splitext(file.name)[1].lower()
        
        if file_extension not in allowed_extensions:
            return {
                "success": False, 
                "error": "Invalid file type. Please upload CSV or Excel files only.",
                "errorType": "INVALID_FILE_TYPE"
            }
        
        # Generate unique filename for temporary storage
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        uploads_dir = os.path.join("uploads")
        file_path = os.path.join(uploads_dir, unique_filename)
        
        # Ensure uploads directory exists
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Save file temporarily
        with open(file_path, "wb") as f:
            for chunk in file.chunks():
                f.write(chunk)
        
        try:
            # Read the file
            if file_extension in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
            else:
                df = pd.read_csv(file_path)
            
            # Validate required columns - be flexible with column names
            possible_sku_columns = ['Child sku', 'child sku', 'Child SKU', 'child_sku']
            possible_store_columns = ['store name', 'Store name', 'Store Name', 'store_name']
            
            sku_column = None
            store_column = None
            
            for col in df.columns:
                if col in possible_sku_columns:
                    sku_column = col
                if col in possible_store_columns:
                    store_column = col
            
            if not sku_column or not store_column:
                return {
                    "success": False,
                    "error": f"Missing required columns. Expected: 'Child sku' and 'store name'. Found columns: {list(df.columns)}",
                    "errorType": "MISSING_COLUMNS"
                }
            
            # Check for empty data
            if df.empty:
                return {
                    "success": False,
                    "error": "File is empty or contains no data rows",
                    "errorType": "EMPTY_FILE"
                }
            
            deletion_summary = {
                "products_deleted": 0,
                "rows_processed": 0,
                "rows_not_found": 0
            }
            
            not_found_items = []
            
            # Process deletion within transaction
            with transaction.atomic():
                for index, row in df.iterrows():
                    child_sku = str(row[sku_column]).strip()
                    store_name = str(row[store_column]).strip()
                    
                    # Skip empty rows
                    if not child_sku or child_sku.lower() in ['nan', 'none', '']:
                        continue
                    if not store_name or store_name.lower() in ['nan', 'none', '']:
                        continue
                    
                    deletion_summary["rows_processed"] += 1
                    
                    # Find products matching this SKU and store name
                    matching_products = Product.objects.filter(
                        marketplace_child_sku=child_sku,
                        store__name=store_name
                    ).select_related('store', 'marketplace', 'vendor')
                    
                    if not matching_products.exists():
                        not_found_items.append(f"Row {index + 1}: SKU '{child_sku}' in store '{store_name}' not found")
                        deletion_summary["rows_not_found"] += 1
                        continue
                    
                    # Delete products (VendorPrice will be automatically deleted due to CASCADE)
                    products_deleted = matching_products.count()
                    matching_products.delete()
                    deletion_summary["products_deleted"] += products_deleted
            
            # Clean up temporary file
            if os.path.exists(file_path):
                os.remove(file_path)
            
            # Prepare response - match frontend expectations
            message = f"Bulk deletion completed. {deletion_summary['products_deleted']} products deleted."
            
            response_data = {
                "success": True,
                "message": message,
                "deletedCount": deletion_summary["products_deleted"],  # Frontend expects this field
                "summary": deletion_summary
            }
            
            # Include warnings if some items weren't found
            if not_found_items:
                response_data["warnings"] = not_found_items[:10]  # Limit to first 10 warnings
                if len(not_found_items) > 10:
                    response_data["warnings"].append(f"... and {len(not_found_items) - 10} more items not found")
            
            return response_data
            
        except Exception as e:
            # Clean up temporary file on error
            if os.path.exists(file_path):
                os.remove(file_path)
            
            return {
                "success": False,
                "error": f"File processing error: {str(e)}",
                "errorType": "FILE_PROCESSING_ERROR"
            }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"Bulk delete failed: {str(e)}",
            "errorType": "BULK_DELETE_ERROR"
        }

@router.get("/export/")
def export_products(request):
    """
    Export all current products as CSV with vendor price and inventory data
    """
    try:
        # Get all products with related data including vendor prices
        products = Product.objects.select_related('vendor', 'marketplace', 'store').prefetch_related('latest_price').all()
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="system_products.csv"'
        
        # Create CSV writer
        writer = csv.writer(response)
        
        # Write headers - added Vendor Price and Vendor Inventory
        headers = [
            'Vendor Name',
            'Vendor ID',
            'Is Variation',
            'Variation ID', 
            'Marketplace Name',
            'Store Name',
            'Marketplace Parent SKU',
            'Marketplace Child SKU',
            'Marketplace ID',
            'Vendor Price',  # New column
            'Vendor Inventory'  # New column
        ]
        writer.writerow(headers)
        
        # Write product data
        for product in products:
            # Get the latest vendor price for this product
            vendor_price = product.latest_price  # Gets the related VendorPrice object
            
            row = [
                product.vendor.name if product.vendor else '',
                product.vendor_sku or '',
                'Yes' if product.variation_id else 'No',
                product.variation_id or '',
                product.marketplace.name if product.marketplace else '',
                product.store.name if product.store else '',
                product.marketplace_parent_sku or '',
                product.marketplace_child_sku or '',
                product.marketplace_external_id or '',  # External marketplace ID
                vendor_price.price if vendor_price and vendor_price.price else '',  # Vendor Price
                vendor_price.stock if vendor_price and vendor_price.stock else '0'  # Vendor Inventory
            ]
            writer.writerow(row)
        
        return response
        
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": f"Export failed: {str(e)}",
            "errorType": "EXPORT_ERROR"
        })

# ========== SCRAPING FUNCTIONALITY ==========

def get_random_headers() -> Dict[str, str]:
    """Generate random headers for HTTP requests."""
    return {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-US,en;q=0.9',
        'cache-control': 'max-age=0',
        'priority': 'u=0, i',
        'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': random.choice(USER_AGENTS),
        'referer': 'https://www.ebay.ca/',
        'dnt': '1',
        'connection': 'keep-alive',
    }

async def is_blocked_content(content: str) -> bool:
    """Detect if response indicates blocking."""
    lower_content = content.lower()
    block_indicators = [
        'captcha', 'recaptcha', 'verify you are human',
        'robot check', 'security page', 'access denied',
        'you have been blocked', 'suspicious activity',
        'please enable cookies', 'browser check', 'just a moment',
        'checking your browser', 'ddos protection', 'cloudflare'
    ]
    return any(indicator in lower_content for indicator in block_indicators)

def extract_product_data(soup: BeautifulSoup) -> Dict[str, Any]:
    """Extract all product data from BeautifulSoup object."""
    try:
        # Title extraction
        title_element = soup.select_one(SELECTORS['title'])
        title = title_element.get_text(strip=True) if title_element else "Title not found"
        
        # Price extraction
        price_element = soup.select_one(SELECTORS['price'])
        price = price_element.get_text(strip=True) if price_element else None
        
        # Stock extraction
        stock_element = soup.select_one(SELECTORS['stock'])
        stock = stock_element.get_text(strip=True) if stock_element else None
        
        # Additional data extraction
        message_element = soup.select_one(SELECTORS['message'])
        quantity_info = get_quantity_from_source(str(soup))
        
        # Use message content as quantity if quantity info not found
        if quantity_info == "Quantity info not found" and message_element:
            quantity_info = message_element.get_text(strip=True)
        
        return {
            'success': True,
            'title': title,
            'price': price,
            'stock': stock,
            'quantity': quantity_info,
            'ended_listings': get_ended_listings(soup),
            'seller_away': get_seller_away(soup),
            'shipping_info': get_shipping_info(soup),
            'handling_time': get_handling_time(str(soup)),
            'category_hierarchy': get_category_hierarchy(soup),
            'variation_count': get_variation_count(soup),
            'error_status': ""
        }
    
    except Exception as e:
        logger.error(f"Error extracting product data: {e}")
        return {
            'success': False,
            'error_status': f'Data extraction error: {str(e)}'
        }

def get_quantity_from_source(page_source: str) -> str:
    """Extract quantity information from page source."""
    match = QUANTITY_PATTERN.search(page_source)
    if match:
        min_value = match.group(1)
        max_value = match.group(2)
        return f"Min: {min_value}, Max: {max_value}"
    return "Quantity info not found"

def get_ended_listings(soup: BeautifulSoup) -> str:
    """Extract ended listings status."""
    status_element = soup.select_one(SELECTORS['status_message'])
    return status_element.get_text(strip=True) if status_element else ""

def get_seller_away(soup: BeautifulSoup) -> str:
    """Extract seller away message."""
    seller_away_element = soup.select_one(SELECTORS['seller_away'])
    return seller_away_element.get_text(strip=True) if seller_away_element else ""

def get_shipping_info(soup: BeautifulSoup) -> str:
    """Extract shipping information."""
    shipping_element = soup.select_one(SELECTORS['shipping'])
    return shipping_element.get_text(strip=True) if shipping_element else "No shipping info"

def get_handling_time(page_source: str) -> str:
    """Extract handling time from page source."""
    match = HANDLING_PATTERN.search(page_source)
    if match:
        return f"Will usually ship within {match.group()}"
    return "Handling time info not found"

def get_category_hierarchy(soup: BeautifulSoup) -> str:
    """Extract category hierarchy from breadcrumbs."""
    breadcrumb_elements = soup.select(SELECTORS['breadcrumb'])
    categories = [elem.get_text(strip=True) for elem in breadcrumb_elements]
    return " > ".join(categories) if categories else "Category not found"

def get_variation_count(soup: BeautifulSoup) -> int:
    """Count product variations."""
    select_boxes = soup.select(SELECTORS['select_boxes'])
    return len(select_boxes)

async def scrape_single_product(
    product: Product, 
    session: aiohttp.ClientSession,
    retries: int = 0
) -> Dict[str, Any]:
    """Scrape a single product's data from eBay."""
    # Generate eBay URL
    item_number = str(product.vendor_sku).split('.')[0]
    url = f"https://www.ebay.ca/itm/{item_number}"
    
    try:
        # Apply delays and backoff
        if retries > 0:
            await asyncio.sleep(2 ** retries + random.uniform(1, 3))
        else:
            await asyncio.sleep(random.uniform(2.5, 6.5))

        headers = get_random_headers()
        
        async with session.get(url, timeout=TIMEOUT, headers=headers) as response:
            content = await response.text()

            # Check for blocking
            if await is_blocked_content(content):
                if retries < RETRY_LIMIT:
                    logger.warning(f"Blocked page for {product.vendor_sku}, retrying...")
                    return await scrape_single_product(product, session, retries + 1)
                else:
                    return {
                        'product_id': product.id,
                        'success': False,
                        'error_status': 'Blocked by security (CAPTCHA/Robot check)'
                    }

            if response.status != 200:
                if retries < RETRY_LIMIT:
                    logger.warning(f"HTTP {response.status} for {product.vendor_sku}, retrying...")
                    await asyncio.sleep(2 ** retries + random.uniform(1, 3))
                    return await scrape_single_product(product, session, retries + 1)
                else:
                    return {
                        'product_id': product.id,
                        'success': False,
                        'error_status': f'HTTP {response.status}'
                    }

            soup = BeautifulSoup(content, 'lxml')
            result = extract_product_data(soup)
            result['product_id'] = product.id
            result['url'] = url
            
            return result

    except asyncio.TimeoutError:
        if retries < RETRY_LIMIT:
            logger.warning(f"Timeout for {product.vendor_sku}, retrying...")
            return await scrape_single_product(product, session, retries + 1)
        else:
            return {
                'product_id': product.id,
                'success': False,
                'error_status': 'Request timed out'
            }
    
    except Exception as e:
        if retries < RETRY_LIMIT:
            logger.warning(f"Error scraping {product.vendor_sku}: {e}, retrying...")
            await asyncio.sleep(2 ** retries + random.uniform(1, 4))
            return await scrape_single_product(product, session, retries + 1)
        else:
            logger.error(f"Final error scraping {product.vendor_sku}: {e}")
            return {
                'product_id': product.id,
                'success': False,
                'error_status': f'Exception: {str(e)}'
            }

async def process_products_batch(
    products_batch: List[Product],
    session: aiohttp.ClientSession
) -> List[Dict[str, Any]]:
    """Process a batch of products concurrently."""
    tasks = [
        scrape_single_product(product, session)
        for product in products_batch
    ]
    return await asyncio.gather(*tasks)

# eBayAU Scraping Functions
async def scrape_single_ebayau_product(product: Product, session: aiohttp.ClientSession, retries: int = 0) -> Dict[str, Any]:
    """Scrape a single eBayAU product with 3 retry attempts"""
    
    # Generate eBayAU URL with cleaned vendor_sku
    item_number = str(product.vendor_sku).split('.')[0]  # Clean vendor_sku
    url = f"https://www.ebay.com.au/itm/{item_number}"
    
    logger.debug(f"Scraping product {product.id} (SKU: {product.vendor_sku}) - URL: {url} - Retry: {retries}")
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Referer": "https://www.ebay.com.au/",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0",
        "Pragma": "no-cache",
        "DNT": "1"
    }
    
    error_output = ""
    product_details = {}
    
    try:
        # Apply delays and backoff
        if retries > 0:
            delay = 2 ** retries + random.uniform(1, 3)
            logger.debug(f"Retry {retries} - waiting {delay:.2f} seconds")
            await asyncio.sleep(delay)
        else:
            delay = random.uniform(2.5, 6.5)
            logger.debug(f"Initial request - waiting {delay:.2f} seconds")
            await asyncio.sleep(delay)
        
        # Change domain for request (as per your script)
        parsed_url = urlparse(url)
        modified_netloc = parsed_url.netloc.replace("ebay.com.au", "ebay.ca")
        modified_url = parsed_url._replace(netloc=modified_netloc).geturl()
        
        logger.debug(f"Making request to modified URL: {modified_url}")
        
        async with session.get(modified_url, timeout=EBAYAU_TIMEOUT, headers=headers) as response:
            logger.debug(f"Response status: {response.status} for product {product.id}")
            
            content = await response.text(errors='ignore')
            soup = BeautifulSoup(content, 'lxml')
            
            # Check for specific error
            specific_error_element = soup.select_one(EBAYAU_SELECTORS['specific_error_header'])
            if specific_error_element:
                error_output = specific_error_element.get_text(strip=True)
                logger.debug(f"Specific error found for product {product.id}: {error_output}")
            elif response.status != 200:
                error_output = f"Failed to retrieve: Status {response.status}"
                logger.debug(f"HTTP error for product {product.id}: {error_output}")
            
            # Extract data if successful
            if response.status == 200 and not error_output:
                logger.debug(f"Extracting data for product {product.id}")
                product_details = parse_ebayau_product_details_from_soup(soup, url)
                select_boxes = soup.select(EBAYAU_SELECTORS['select_boxes'])
                product_details['count'] = len(select_boxes)
                
                logger.debug(f"Product {product.id} data extracted: {product_details}")
    
    except asyncio.TimeoutError:
        if retries < EBAYAU_RETRY_LIMIT:
            logger.warning(f"Timeout for product {product.id} (SKU: {product.vendor_sku}), retry {retries + 1}/{EBAYAU_RETRY_LIMIT}")
            return await scrape_single_ebayau_product(product, session, retries + 1)
        else:
            error_output = f"Request timed out for {url}"
            logger.error(f"Final timeout for product {product.id} (SKU: {product.vendor_sku})")
    
    except aiohttp.ClientError as e:
        if retries < EBAYAU_RETRY_LIMIT:
            logger.warning(f"Client error for product {product.id} (SKU: {product.vendor_sku}): {e}, retry {retries + 1}/{EBAYAU_RETRY_LIMIT}")
            await asyncio.sleep(2 ** retries + random.uniform(1, 4))
            return await scrape_single_ebayau_product(product, session, retries + 1)
        else:
            error_output = f"Client error for {url}: {str(e)}"
            logger.error(f"Final client error for product {product.id} (SKU: {product.vendor_sku}): {e}")
    
    except Exception as e:
        if retries < EBAYAU_RETRY_LIMIT:
            logger.warning(f"Error scraping product {product.id} (SKU: {product.vendor_sku}): {e}, retry {retries + 1}/{EBAYAU_RETRY_LIMIT}")
            await asyncio.sleep(2 ** retries + random.uniform(1, 4))
            return await scrape_single_ebayau_product(product, session, retries + 1)
        else:
            logger.error(f"Final error scraping product {product.id} (SKU: {product.vendor_sku}): {e}")
            error_output = f"An unexpected error occurred for {url}: {str(e)}"
    
    result = {
        'product_id': product.id,
        'vendor_sku': product.vendor_sku,
        'url': url,
        'success': not bool(error_output),
        'error_status': error_output,
        **product_details
    }
    
    logger.debug(f"Final result for product {product.id}: success={result['success']}, error={error_output}")
    
    return result

async def process_ebayau_batch(products_batch: List[Product], session: aiohttp.ClientSession) -> List[Dict[str, Any]]:
    """Process a batch of eBayAU products concurrently"""
    tasks = [
        scrape_single_ebayau_product(product, session)
        for product in products_batch
    ]
    return await asyncio.gather(*tasks)

@transaction.atomic
def save_ebayau_scraping_results(results: List[Dict[str, Any]]) -> List[int]:
    """Save eBayAU results and return product IDs that need rescraping"""
    db_logger.info(f"=== DATABASE SAVE START ===")
    db_logger.info(f"Total results to save: {len(results)}")
    
    rescrape_product_ids = []
    scrape_time = timezone.now()  # UTC timestamp
    
    db_logger.info(f"Scrape time: {scrape_time}")
    
    for i, result in enumerate(results):
        try:
            product_id = result.get('product_id')
            db_logger.info(f"Processing result {i+1}/{len(results)} - Product ID: {product_id}")
            
            product = Product.objects.get(id=product_id)
            db_logger.info(f"Found product: {product.vendor_sku} (Vendor: {product.vendor.name})")
            
            # Apply business rules
            db_logger.info(f"Applying business rules for product {product_id}")
            processed_data = eBayAUBusinessRules.process_scraped_data(result)
            db_logger.info(f"Business rules processed - needs_rescrape: {processed_data['needs_rescrape']}")
            db_logger.info(f"Final price: {processed_data['final_price']}")
            db_logger.info(f"Final inventory: {processed_data['final_inventory']}")
            
            # Create Scrape record
            db_logger.info(f"Creating Scrape record for product {product_id}")
            scrape = Scrape.objects.create(
                product=product,
                scrape_time=scrape_time,
                raw_response=result,
                error_code=processed_data['error_details'],
                raw_price=processed_data['raw_price'],
                raw_shipping=processed_data['raw_shipping'],
                raw_quantity=processed_data['raw_quantity'],
                raw_handling_time=processed_data['raw_handling_time'],
                raw_seller_away=processed_data['raw_seller_away'],
                raw_ended_listings=processed_data['raw_ended_listings'],
                calculated_shipping_price=processed_data['calculated_shipping_price'],
                final_price=processed_data['final_price'],
                final_inventory=processed_data['final_inventory'],
                needs_rescrape=processed_data['needs_rescrape'],
                error_details=processed_data['error_details']
            )
            db_logger.info(f"Scrape record created with ID: {scrape.id}")
            
            # Update VendorPrice with final calculated values
            db_logger.info(f"Updating VendorPrice for product {product_id}")
            db_logger.info(f"Price: {processed_data['final_price']}, Stock: {processed_data['final_inventory']}")
            
            vendor_price, created = VendorPrice.objects.update_or_create(
                product=product,
                defaults={
                    'price': processed_data['final_price'],
                    'stock': processed_data['final_inventory'],  # Using final_inventory as stock
                    'error_code': processed_data['error_details'],
                    'scraped_at': scrape_time
                }
            )
            
            if created:
                db_logger.info(f"VendorPrice record CREATED for product {product_id}")
            else:
                db_logger.info(f"VendorPrice record UPDATED for product {product_id}")
            
            db_logger.info(f"VendorPrice ID: {vendor_price.id}")
            
            # Track products that need rescraping (return actual product IDs)
            if processed_data['needs_rescrape']:
                rescrape_product_ids.append(product.id)
                db_logger.info(f"Product {product_id} marked for rescraping")
                
        except Product.DoesNotExist:
            db_logger.error(f"Product {result.get('product_id')} not found in database")
        except Exception as e:
            db_logger.error(f"Error saving result for product {result.get('product_id')}: {e}")
            db_logger.error(f"Error type: {type(e)}")
            db_logger.error("Error traceback: ", exc_info=True)
    
    db_logger.info(f"=== DATABASE SAVE COMPLETE ===")
    db_logger.info(f"Total products processed: {len(results)}")
    db_logger.info(f"Products needing rescrape: {len(rescrape_product_ids)}")
    db_logger.info(f"Rescrape product IDs: {rescrape_product_ids}")
    
    return rescrape_product_ids

async def run_ebayau_scraping_job(session_id: str):
    """Complete eBayAU scraping job with vendor name filtering and SKU dedupe"""
    start_time = timezone.now()
    
    logger.info(f"=== EBAYAU SCRAPING JOB START ===")
    logger.info(f"Session ID: {session_id}")
    logger.info(f"Start time: {start_time}")
    
    try:
        # Get products with eBayAU vendor name variations (async)
        logger.info("Fetching eBayAU products from database...")
        products = await get_ebayau_products()
        total_products = len(products)  # total rows to update
        
        logger.info(f"Found {total_products} products with eBayAU vendor names")
        logger.info(f"Vendor names being processed: {eBayAUBusinessRules.EBAYAU_VENDOR_VARIATIONS}")
        
        if total_products == 0:
            logger.info("No products found with eBayAU vendor name variations")
            return
        
        # Group by (vendor_id, normalized vendor_sku) to dedupe
        rep_products, rep_to_ids = build_vendor_sku_groups(products)
        total_unique = len(rep_products)
        logger.info(f"Deduped set: {total_unique} unique SKUs from {total_products} products")
        logger.info(f"Starting eBayAU scraping job {session_id}")
        
        # Configure session
        logger.info("Configuring aiohttp session...")
        connector = aiohttp.TCPConnector(limit=EBAYAU_MAX_CONCURRENT_REQUESTS, force_close=True)
        async with aiohttp.ClientSession(
            connector=connector, 
            timeout=EBAYAU_TIMEOUT, 
            cookies=EBAYAU_COOKIES
        ) as session:
            
            total_processed = 0
            all_rescrape_ids = []
            
            # Process representative products in batches
            total_batches = (total_unique + EBAYAU_BATCH_SIZE - 1) // EBAYAU_BATCH_SIZE
            logger.info(f"Processing in {total_batches} batches of {EBAYAU_BATCH_SIZE} representatives each")
            
            for i in range(0, total_unique, EBAYAU_BATCH_SIZE):
                batch_num = i // EBAYAU_BATCH_SIZE + 1
                reps_batch = rep_products[i:i + EBAYAU_BATCH_SIZE]
                
                logger.info(f"Processing batch {batch_num}/{total_batches} ({len(reps_batch)} reps)")
                
                # Scrape reps
                logger.info(f"Starting scraping for batch {batch_num}")
                batch_results = await process_ebayau_batch(reps_batch, session)
                logger.info(f"Batch {batch_num} scraping completed, {len(batch_results)} rep results received")
                
                # Fan-out results to all product IDs sharing the same vendor_sku+vendor
                fanout_results = []
                for r in batch_results:
                    targets = rep_to_ids.get(r['product_id'], [r['product_id']])
                    for pid in targets:
                        nr = dict(r)
                        nr['product_id'] = pid
                        fanout_results.append(nr)

                # Log sample results
                successful_results = [r for r in fanout_results if r.get('success')]
                failed_results = [r for r in fanout_results if not r.get('success')]
                logger.info(f"Batch {batch_num} (expanded) - Successful: {len(successful_results)}, Failed: {len(failed_results)}")
                if failed_results:
                    logger.info(f"Sample failed results from batch {batch_num}:")
                    for j, failed in enumerate(failed_results[:3]):
                        logger.info(f"  Failed {j+1}: Product {failed.get('product_id')} - {failed.get('error_status')}")
                
                # Save expanded results
                logger.info(f"Saving batch {batch_num} expanded results to database...")
                rescrape_ids = await sync_to_async(save_ebayau_scraping_results)(fanout_results)
                logger.info(f"Batch {batch_num} saved - {len(rescrape_ids)} products need rescraping")
                
                all_rescrape_ids.extend(rescrape_ids)
                total_processed += len(fanout_results)
                logger.info(f"Progress: {total_processed}/{total_products} ({(total_processed/total_products)*100:.1f}%)")
            
            # Completion
            duration = timezone.now() - start_time
            successful_scrapes = total_products - len(all_rescrape_ids)
            
            logger.info(f"=== EBAYAU SCRAPING JOB COMPLETE ===")
            logger.info(f"Session ID: {session_id}")
            logger.info(f"Total product-rows updated: {total_products}")
            logger.info(f"Unique SKUs scraped: {total_unique}")
            logger.info(f"Successful scrapes: {successful_scrapes}")
            logger.info(f"Failed scrapes: {len(all_rescrape_ids)}")
            logger.info(f"Success rate: {(successful_scrapes/total_products)*100:.1f}%")
            logger.info(f"Duration: {duration}")
            logger.info(f"Products needing rescrape: {len(all_rescrape_ids)}")
            
            # Trigger webhook if needed
            if all_rescrape_ids:
                logger.info(f"=== TRIGGERING N8N WEBHOOK ===")
                logger.info(f"Products needing rescrape: {len(all_rescrape_ids)}")
                logger.info(f"Rescrape product IDs: {all_rescrape_ids[:10]}...")
                
                webhook_success = await trigger_n8n_rescrape_webhook(all_rescrape_ids, session_id)
                if webhook_success:
                    logger.info("SUCCESS: n8n webhook triggered successfully")
                else:
                    logger.error("ERROR: Failed to trigger n8n webhook - rescraping may not happen automatically")
            else:
                logger.info("No products need rescraping, n8n webhook not triggered")
            
            # CSV + email (unchanged, but stats reflect fan-out)
            try:
                logger.info("=== GENERATING SYSTEM PRODUCTS CSV ===")
                csv_file_path = generate_system_products_csv()
                scraping_stats = {
                    'total_products': total_products,
                    'successful_scrapes': successful_scrapes,
                    'failed_scrapes': len(all_rescrape_ids),
                    'success_rate': (successful_scrapes/total_products)*100 if total_products > 0 else 0,
                    'duration': duration
                }
                logger.info("=== SENDING SCRAPING COMPLETION EMAIL ===")
                email_success = await asyncio.to_thread(
                    send_scraping_complete_email, session_id, scraping_stats, csv_file_path
                )
                if email_success:
                    logger.info("SUCCESS: Scraping completion email sent successfully")
                else:
                    logger.error("ERROR: Failed to send scraping completion email")
            except Exception as email_error:
                logger.error(f"Error sending scraping completion email: {email_error}")
            
    except Exception as e:
        logger.error(f"=== EBAYAU SCRAPING JOB ERROR ===")
        logger.error(f"Session ID: {session_id}")
        logger.error(f"Error: {e}")
        logger.error(f"Error type: {type(e)}")
        logger.error("Error traceback: ", exc_info=True)

async def run_ebayau_rescraping_job(session_id: str, product_ids: List[int]):
    """Rescrape specific eBayAU products with SKU dedupe"""
    try:
        # Get products by IDs (async)
        products = await get_products_by_ids(product_ids)
        
        # Deduplicate by (vendor_id, normalized vendor_sku)
        rep_products, rep_to_ids = build_vendor_sku_groups(products)

        connector = aiohttp.TCPConnector(limit=EBAYAU_MAX_CONCURRENT_REQUESTS, force_close=True)
        async with aiohttp.ClientSession(
            connector=connector, 
            timeout=EBAYAU_TIMEOUT, 
            cookies=EBAYAU_COOKIES
        ) as session:
            
            # Process in smaller batches for rescraping
            batch_size = 10
            total_processed = 0
            final_rescrape_ids = []
            
            for i in range(0, len(rep_products), batch_size):
                reps_batch = rep_products[i:i + batch_size]
                batch_results = await process_ebayau_batch(reps_batch, session)

                # Fan-out results to all duplicates
                fanout_results = []
                for r in batch_results:
                    targets = rep_to_ids.get(r['product_id'], [r['product_id']])
                    for pid in targets:
                        nr = dict(r)
                        nr['product_id'] = pid
                        fanout_results.append(nr)

                rescrape_ids = await sync_to_async(save_ebayau_scraping_results)(fanout_results)
                final_rescrape_ids.extend(rescrape_ids)
                total_processed += len(fanout_results)
            
            logger.info(f"Completed eBayAU rescraping job {session_id}: {total_processed} processed, {len(final_rescrape_ids)} still need rescraping")
            
    except Exception as e:
        logger.error(f"Error in eBayAU rescraping job {session_id}: {e}")

@transaction.atomic
def save_scraping_results(results: List[Dict[str, Any]]) -> None:
    """Save scraping results to database efficiently."""
    try:
        # Get timezone for Pakistan time
        tz = pytz.timezone('Asia/Karachi')
        scrape_time = datetime.now(tz)
        
        # Prepare batch updates
        vendor_price_updates = []
        scrape_records = []
        
        for result in results:
            try:
                product = Product.objects.get(id=result['product_id'])
                
                # Parse price and stock
                parsed_price = None
                parsed_stock = None
                
                if result.get('success'):
                    parsed_price = parse_price_to_decimal(result.get('price'))
                    parsed_stock = parse_stock_to_int(result.get('stock'))
                
                # Update VendorPrice
                vendor_price, created = VendorPrice.objects.get_or_create(
                    product=product,
                    defaults={
                        'price': parsed_price,
                        'stock': parsed_stock,
                        'error_code': result.get('error_status', ''),
                        'scraped_at': scrape_time
                    }
                )
                
                if not created:
                    vendor_price.price = parsed_price
                    vendor_price.stock = parsed_stock
                    vendor_price.error_code = result.get('error_status', '')
                    vendor_price.scraped_at = scrape_time
                    vendor_price_updates.append(vendor_price)
                
                # Create Scrape record
                scrape_record = Scrape(
                    product=product,
                    scrape_time=scrape_time,
                    stock=parsed_stock,
                    error_code=result.get('error_status', ''),
                    raw_response=result
                )
                scrape_records.append(scrape_record)
                
            except Product.DoesNotExist:
                logger.error(f"Product {result['product_id']} not found")
            except Exception as e:
                logger.error(f"Error saving result for product {result.get('product_id')}: {e}")
        
        # Batch save operations
        if vendor_price_updates:
            VendorPrice.objects.bulk_update(
                vendor_price_updates,
                ['price', 'stock', 'error_code', 'scraped_at'],
                batch_size=100
            )
        
        if scrape_records:
            Scrape.objects.bulk_create(scrape_records, batch_size=100)
            
        logger.info(f"Saved {len(scrape_records)} scrape results to database")
        
    except Exception as e:
        logger.error(f"Error saving scraping results: {e}")
        raise

async def run_complete_scraping_job(session_id: str) -> Dict[str, Any]:
    """Execute the complete scraping job asynchronously."""
    start_time = datetime.now()
    logger.info(f"Starting scraping job {session_id}")
    
    try:
        # Get eBayUS marketplace
        ebay_marketplace = Marketplace.objects.filter(
            Q(code="eBayUS") | Q(name="eBayUS")
        ).first()
        
        if not ebay_marketplace:
            return {
                'success': False,
                'error': 'eBayUS marketplace not found in database',
                'session_id': session_id,
                'total_products': 0
            }
        
        # Get products to scrape
        products = Product.objects.filter(
            marketplace=ebay_marketplace,
            store__is_active=True
        ).select_related('vendor', 'marketplace', 'store', 'upload')
        
        # Validate eBay item numbers
        valid_products = []
        for product in products:
            is_valid, _ = validate_ebay_item_number(product.vendor_sku)
            if is_valid:
                valid_products.append(product)
        
        if not valid_products:
            return {
                'success': False,
                'error': 'No valid eBayUS products found to scrape',
                'session_id': session_id,
                'total_products': 0
            }
        
        # Setup HTTP session with proper connector
        connector = aiohttp.TCPConnector(
            limit_per_host=MAX_CONCURRENT_REQUESTS,
            ssl=False
        )
        
        all_results = []
        
        async with aiohttp.ClientSession(
            connector=connector,
            timeout=TIMEOUT
        ) as session:
            # Process in batches
            for i in range(0, len(valid_products), BATCH_SIZE):
                batch = valid_products[i:i + BATCH_SIZE]
                batch_results = await process_products_batch(batch, session)
                all_results.extend(batch_results)
                
                logger.info(f"Processed batch {i//BATCH_SIZE + 1}, {len(all_results)}/{len(valid_products)} products")
        
        # Save results to database
        save_scraping_results(all_results)
        
        # Update store last_scrape_time
        Store.objects.filter(marketplace=ebay_marketplace, is_active=True).update(
            last_scrape_time=datetime.now()
        )
        
        # Create error log Excel file
        error_log_file = create_error_log_excel(all_results, session_id)
        
        # Calculate statistics
        successful_scrapes = sum(1 for r in all_results if r.get('success'))
        failed_scrapes = len(all_results) - successful_scrapes
        duration = datetime.now() - start_time
        
        result = {
            'success': True,
            'session_id': session_id,
            'total_products': len(valid_products),
            'successful_scrapes': successful_scrapes,
            'failed_scrapes': failed_scrapes,
            'duration_minutes': int(duration.total_seconds() // 60),
            'error_log_file': error_log_file
        }
        
        logger.info(f"Completed scraping job {session_id}: {successful_scrapes}/{len(valid_products)} successful")
        
        # Generate CSV and send completion email
        try:
            logger.info("=== GENERATING SYSTEM PRODUCTS CSV ===")
            csv_file_path = generate_system_products_csv()
            
            # Prepare scraping statistics
            scraping_stats = {
                'total_products': len(valid_products),
                'successful_scrapes': successful_scrapes,
                'failed_scrapes': failed_scrapes,
                'success_rate': (successful_scrapes/len(valid_products))*100 if len(valid_products) > 0 else 0,
                'duration': duration
            }
            
            logger.info("=== SENDING SCRAPING COMPLETION EMAIL ===")
            email_success = await asyncio.to_thread(
                send_scraping_complete_email, session_id, scraping_stats, csv_file_path
            )
            
            if email_success:
                logger.info("SUCCESS: Scraping completion email sent successfully")
            else:
                logger.error("ERROR: Failed to send scraping completion email")
                
        except Exception as email_error:
            logger.error(f"Error sending scraping completion email: {email_error}")
        
        return result
        
    except Exception as e:
        logger.error(f"Error in scraping job {session_id}: {e}")
        return {
            'success': False,
            'error': str(e),
            'session_id': session_id,
            'total_products': 0
        }

@router.post("/scrape/")
async def scrape_products(request):
    """
    Async API to scrape all eBayUS products and update prices/inventory.
    
    This endpoint starts an asynchronous scraping job for all valid eBayUS products.
    It returns immediately with job details, and the actual scraping runs in the background.
    """
    try:
        # Generate session ID
        session_id = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        # Get quick product count for response
        ebay_marketplace = Marketplace.objects.filter(
            Q(code="eBayUS") | Q(name="eBayUS")
        ).first()
        
        if not ebay_marketplace:
            return {
                "success": False,
                "error": "eBayUS marketplace not found in database",
                "products_queued": 0
            }
        
        # Quick count of products
        total_products = Product.objects.filter(marketplace=ebay_marketplace, store__is_active=True).count()
        
        # Start scraping job in background
        asyncio.create_task(run_complete_scraping_job(session_id))
        
        return {
            "success": True,
            "message": "eBayUS scraping started successfully",
            "session_id": session_id,
            "products_queued": total_products,
            "estimated_duration": f"{total_products * 5 // 60} minutes",
            "status": "Scraping job started in background"
        }
        
    except Exception as e:
        logger.error(f"Error starting scraping job: {e}")
        return {
            "success": False,
            "error": f"Failed to start scraping: {str(e)}",
            "products_queued": 0
        }

def start_detached_scrape(session_id: str) -> str:
    """Start the eBayAU scraping job as a detached subprocess and return log file path."""
    uploads_dir = os.path.join("uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    log_path = os.path.join(uploads_dir, f"scrape_{session_id}.log")
    with open(log_path, "ab", buffering=0) as lf:
        subprocess.Popen(
            [sys.executable, "manage.py", "scrape_ebayau_job", "--session", session_id],
            cwd=os.getcwd(),
            stdout=lf,
            stderr=lf,
            start_new_session=True
        )
    return log_path

@router.post("/scrape-ebayau/")
async def scrape_ebayau_products(request):
    """
    Async API to scrape all products with eBayAU vendor name variations and update prices/inventory.
    Applies specific business rules for eBayAU marketplace.
    """
    logger.info(f"=== EBAYAU SCRAPING API ENDPOINT CALLED ===")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request headers: {dict(request.headers)}")
    
    try:
        session_id = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        logger.info(f"Generated session ID: {session_id}")
        
        # Get products with eBayAU vendor name variations (async)
        logger.info("Counting eBayAU products...")
        total_products = await get_ebayau_products_count()
        logger.info(f"Found {total_products} products with eBayAU vendor names")
        
        if total_products == 0:
            logger.warning("No products found with eBayAU vendor name variations")
            return {
                "success": False,
                "error": "No products found with eBayAU vendor name variations",
                "products_queued": 0
            }
        
        # Start eBayAU scraping job as detached process
        logger.info(f"Starting eBayAU scraping job with session ID: {session_id}")
        log_path = start_detached_scrape(session_id)
        logger.info("eBayAU scraping job subprocess started successfully")
        
        estimated_duration = f"{total_products * 6 // 60} minutes"
        logger.info(f"Estimated duration: {estimated_duration}")
        
        return {
            "success": True,
            "message": "eBayAU scraping started successfully",
            "session_id": session_id,
            "products_queued": total_products,
            "estimated_duration": estimated_duration,
            "status": "Scraping job started in background",
            "log_file": log_path
        }
        
    except Exception as e:
        logger.error(f"Error starting scraping job: {e}")
        return {
            "success": False,
            "error": f"Failed to start scraping: {str(e)}",
            "products_queued": 0
        }

@router.post("/rescrape-ebayau/")
async def rescrape_failed_ebayau_products(request):
    """
    Rescrape eBayAU products that failed with 503 errors after 3 retry attempts.
    Returns list of actual product IDs that still need rescraping.
    """
    try:
        # Get products that need rescraping (async)
        rescrape_products = await get_rescrape_products()
        
        if not rescrape_products:
            return {
                "success": True,
                "message": "No products need rescraping",
                "rescrape_product_ids": [],
                "rescrape_vendor_skus": []
            }
        
        # Get actual product IDs and vendor SKUs
        product_ids = [p.id for p in rescrape_products]
        vendor_skus = [p.vendor_sku for p in rescrape_products]
        
        # Start rescraping job
        session_id = datetime.now().strftime('%Y-%m-%d_%H-%M-%S_rescrape')
        asyncio.create_task(run_ebayau_rescraping_job(session_id, product_ids))
        
        return {
            "success": True,
            "message": f"Rescraping started for {len(product_ids)} products",
            "session_id": session_id,
            "rescrape_product_ids": product_ids,  # Actual product IDs
            "rescrape_vendor_skus": vendor_skus,
            "products_queued": len(product_ids)
        }
        
    except Exception as e:
        logger.error(f"Error starting eBayAU rescraping: {e}")
        return {
            "success": False,
            "error": str(e),
            "rescrape_product_ids": []
        }

@router.post("/check-rescrape-status/")
async def check_rescrape_status(request):
    """
    Check the status of a rescraping job by session_id.
    Returns completion status and final rescrape count.
    """
    try:
        session_id = request.POST.get('session_id')
        
        if not session_id:
            return {
                "success": False,
                "error": "session_id is required"
            }
        
        # Check if rescraping is still in progress
        # You can implement a more sophisticated status tracking system
        # For now, we'll check if there are still products needing rescraping
        
        rescrape_products = Product.objects.filter(
            vendor__name__in=eBayAUBusinessRules.EBAYAU_VENDOR_VARIATIONS,
            scrapes__needs_rescrape=True,
            store__is_active=True
        ).distinct()
        
        final_rescrape_count = rescrape_products.count()
        
        # If no products need rescraping, consider it completed
        if final_rescrape_count == 0:
            return {
                "success": True,
                "status": "completed",
                "session_id": session_id,
                "total_processed": 0,
                "final_rescrape_count": 0,
                "duration_minutes": 0
            }
        else:
            return {
                "success": True,
                "status": "in_progress",
                "session_id": session_id,
                "remaining_products": final_rescrape_count
            }
        
    except Exception as e:
        logger.error(f"Error checking rescrape status: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@router.get("/n8n-webhook-status/")
async def get_n8n_webhook_status(request):
    """
    Get n8n webhook configuration status
    """
    return {
        "webhook_url": N8N_WEBHOOK_URL,
        "timeout": N8N_WEBHOOK_TIMEOUT,
        "enabled": bool(N8N_WEBHOOK_URL and N8N_WEBHOOK_URL != 'https://autoecom.wesolucions.com/webhook-test/ebayau-rescrape'),
        "environment_variable": "N8N_WEBHOOK_URL"
    }

@router.get("/upload/{upload_id}")
def get_upload_status(request, upload_id: int):
    """Get status and progress for a single upload"""
    upload = get_object_or_404(Upload, id=upload_id)
    info = {}
    if upload.note:
        try:
            info = json.loads(upload.note)
        except Exception:
            info = {}
    # If processing, augment with file-based progress if available
    if info.get('status') == 'processing':
        p = _read_progress(upload_id)
        if p:
            info['itemsProcessed'] = p.get('itemsProcessed', info.get('itemsProcessed', 0))
            info['totalItems'] = p.get('totalItems', info.get('totalItems', info.get('itemsUploaded', 0)))
    return {
        "success": True,
        "upload": {
            "id": upload.id,
            "date": upload.expires_at.strftime("%Y-%m-%d"),
            "userName": "System",
            "vendorName": info.get("vendorName", "Unknown"),
            "marketplace": info.get("marketplace", "Unknown"),
            "itemsUploaded": info.get("itemsUploaded", 0),
            "itemsAdded": info.get("itemsAdded", 0),
            "status": info.get("status", "pending"),
            "errorLogs": info.get("errorLogs", ""),
            "itemsProcessed": info.get("itemsProcessed", 0),
            "totalItems": info.get("totalItems", info.get("itemsUploaded", 0)),
        }
    }

@sync_to_async
def get_amazonau_products_count():
    return Product.objects.filter(
        vendor__name='AmazonAU',
        store__is_active=True
    ).count()

@sync_to_async
def get_amazonau_products():
    return list(Product.objects.filter(
        vendor__name='AmazonAU',
        store__is_active=True
    ))

async def run_amazonau_scraping_job(session_id: str):
    start_time = timezone.now()
    logger.info(f"=== AMAZONAU SCRAPING JOB START === Session: {session_id}")
    try:
        products = await get_amazonau_products()
        total_products = len(products)
        if total_products == 0:
            logger.info("No AmazonAU products found")
            return

        reps, rep_to_ids = AmazonAUScrapper.build_vendor_groups(products)
        total_unique = len(reps)

        connector = aiohttp.TCPConnector(limit=AmazonAUScrapper.AMAZONAU_MAX_CONCURRENT_REQUESTS, force_close=True)
        async with aiohttp.ClientSession(connector=connector, timeout=AmazonAUScrapper.AMAZONAU_TIMEOUT) as session:
            try:
                ok = await AmazonAUScrapper.setup_location_on_session(session)
                logger.info(f"GLUX location set result: {ok}")
            except Exception as se:
                logger.error(f"GLUX setup error (non-fatal): {se}")
            total_processed = 0
            for i in range(0, total_unique, AmazonAUScrapper.AMAZONAU_BATCH_SIZE):
                reps_batch = reps[i:i + AmazonAUScrapper.AMAZONAU_BATCH_SIZE]
                batch_results = await AmazonAUScrapper.process_batch(reps_batch, session)

                expanded = []
                for r in batch_results:
                    targets = rep_to_ids.get(r['product_id'], [r['product_id']])
                    for pid in targets:
                        nr = dict(r)
                        nr['product_id'] = pid
                        expanded.append(nr)

                await sync_to_async(AmazonAUScrapper.save_results)(expanded)
                total_processed += len(expanded)
                logger.info(f"AmazonAU progress: {total_processed}/{total_products}")

        duration = timezone.now() - start_time
        logger.info(f"=== AMAZONAU SCRAPING JOB COMPLETE === Session: {session_id}, duration: {duration}")
        
        # Generate system products CSV and send completion email
        try:
            csv_file_path = await asyncio.to_thread(generate_system_products_csv)
            scraping_stats = {
                'total_products': total_products,
                'successful_scrapes': total_processed,
                'failed_scrapes': 0 if total_products == 0 else max(0, total_products - total_processed),
                'success_rate': (total_processed/total_products)*100 if total_products > 0 else 0,
                'duration': duration
            }
            email_success = await asyncio.to_thread(
                send_scraping_complete_email, session_id, scraping_stats, csv_file_path
            )
            if email_success:
                logger.info("SUCCESS: AmazonAU scraping completion email sent successfully")
            else:
                logger.error("ERROR: Failed to send AmazonAU scraping completion email")
        except Exception as email_error:
            logger.error(f"Error sending AmazonAU scraping completion email: {email_error}")
    except Exception as e:
        logger.error(f"AmazonAU job error: {e}", exc_info=True)


def start_detached_amazon_scrape(session_id: str) -> str:
    uploads_dir = os.path.join("uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    log_path = os.path.join(uploads_dir, f"scrape_amazon_{session_id}.log")
    with open(log_path, "ab", buffering=0) as lf:
        subprocess.Popen(
            [sys.executable, "manage.py", "scrape_amazonau_job", "--session", session_id],
            cwd=os.getcwd(),
            stdout=lf,
            stderr=lf,
            start_new_session=True
        )
    return log_path


@router.post("/scrape-amazonau/")
async def scrape_amazonau_products(request):
    logger.info("=== AMAZONAU SCRAPING API CALLED ===")
    try:
        session_id = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        total_products = await get_amazonau_products_count()
        if total_products == 0:
            return {"success": False, "error": "No products found for AmazonAU Vendor", "products_queued": 0}
        log_path = start_detached_amazon_scrape(session_id)
        return {
            "success": True,
            "message": "AmazonAU scraping started successfully",
            "session_id": session_id,
            "products_queued": total_products,
            "status": "Scraping job started in background",
            "log_file": log_path
        }
    except Exception as e:
        logger.error(f"Error starting AmazonAU scraping: {e}")
        return {"success": False, "error": f"Failed to start AmazonAU scraping: {str(e)}", "products_queued": 0}

@sync_to_async
def get_costcoau_products_count():
    return Product.objects.filter(
        vendor__name__iexact='CostcoAU',
        store__is_active=True
    ).count()

@sync_to_async
def get_costcoau_products():
    return list(Product.objects.filter(
        vendor__name__iexact='CostcoAU',
        store__is_active=True
    ))

async def run_costcoau_scraping_job(session_id: str):
    start_time = timezone.now()
    logger.info(f"=== COSTCOAU SCRAPING JOB START === Session: {session_id}")
    try:
        products = await get_costcoau_products()
        total_products = len(products)
        if total_products == 0:
            logger.info("No CostcoAU products found")
            return

        reps, rep_to_ids = CostcoAUScrapper.build_vendor_groups(products)
        total_unique = len(reps)

        connector = aiohttp.TCPConnector(limit=CostcoAUScrapper.COSTCOAU_MAX_CONCURRENT_REQUESTS, force_close=True)
        async with aiohttp.ClientSession(connector=connector, timeout=CostcoAUScrapper.COSTCOAU_TIMEOUT) as session:
            total_processed = 0
            for i in range(0, total_unique, CostcoAUScrapper.COSTCOAU_BATCH_SIZE):
                reps_batch = reps[i:i + CostcoAUScrapper.COSTCOAU_BATCH_SIZE]
                batch_results = await CostcoAUScrapper.process_batch(reps_batch, session)

                expanded = []
                for r in batch_results:
                    targets = rep_to_ids.get(r['product_id'], [r['product_id']])
                    for pid in targets:
                        nr = dict(r)
                        nr['product_id'] = pid
                        expanded.append(nr)

                await sync_to_async(CostcoAUScrapper.save_results)(expanded)
                total_processed += len(expanded)
                logger.info(f"CostcoAU progress: {total_processed}/{total_products}")

        duration = timezone.now() - start_time
        logger.info(f"=== COSTCOAU SCRAPING JOB COMPLETE === Session: {session_id}, duration: {duration}")

        # Generate system products CSV and email
        try:
            csv_file_path = generate_system_products_csv()
            scraping_stats = {
                'total_products': total_products,
                'successful_scrapes': total_processed,
                'failed_scrapes': 0 if total_products == 0 else max(0, total_products - total_processed),
                'success_rate': (total_processed/total_products)*100 if total_products > 0 else 0,
                'duration': duration
            }
            email_success = await asyncio.to_thread(
                send_scraping_complete_email, session_id, scraping_stats, csv_file_path
            )
            if email_success:
                logger.info("SUCCESS: CostcoAU scraping completion email sent successfully")
            else:
                logger.error("ERROR: Failed to send CostcoAU scraping completion email")
        except Exception as email_error:
            logger.error(f"Error sending CostcoAU scraping completion email: {email_error}")

    except Exception as e:
        logger.error(f"CostcoAU job error: {e}", exc_info=True)


def start_detached_costco_scrape(session_id: str) -> str:
    uploads_dir = os.path.join("uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    log_path = os.path.join(uploads_dir, f"scrape_costco_{session_id}.log")
    with open(log_path, "ab", buffering=0) as lf:
        subprocess.Popen(
            [sys.executable, "manage.py", "scrape_costcoau_job", "--session", session_id],
            cwd=os.getcwd(),
            stdout=lf,
            stderr=lf,
            start_new_session=True
        )
    return log_path


@router.post("/scrape-costcoau/")
async def scrape_costcoau_products(request):
    logger.info("=== COSTCOAU SCRAPING API CALLED ===")
    try:
        session_id = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        total_products = await get_costcoau_products_count()
        if total_products == 0:
            return {"success": False, "error": "No CostcoAU products found", "products_queued": 0}
        log_path = start_detached_costco_scrape(session_id)
        return {
            "success": True,
            "message": "CostcoAU scraping started successfully",
            "session_id": session_id,
            "products_queued": total_products,
            "status": "Scraping job started in background",
            "log_file": log_path
        }
    except Exception as e:
        logger.error(f"Error starting CostcoAU scraping: {e}")
        return {"success": False, "error": f"Failed to start CostcoAU scraping: {str(e)}", "products_queued": 0}
